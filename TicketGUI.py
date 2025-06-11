import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk, filedialog
import csv
import json
import google.generativeai as genai
import random
import re
import os
import threading
import glob
import unicodedata
try:
    import openpyxl
except ImportError:
    messagebox.showerror("Falta openpyxl", "Debes instalar openpyxl: pip install openpyxl")
    raise

# ===================== CONFIGURACIÓN =====================
DEFAULT_EXCEL_FILE = "Tickets_1.xlsx"
DEFAULT_SAVE_FILE = "tickets_generados.csv"
DEFAULT_NUM_EJEMPLOS = 5
GEMINI_API_KEY = "AIzaSyCzmvT35IUVspsf7F8HPAd1GMyltP0hH3A"
MODEL_NAME = "gemini-2.0-flash"

# ===================== LÓGICA DE IA Y ARCHIVOS =====================
def validar_api_key_y_modelo():
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel(MODEL_NAME)
        # Prueba rápida de modelo
        _ = model.generate_content("Test", generation_config={"max_output_tokens": 5})
        return True, ""
    except Exception as e:
        return False, str(e)

def leer_agentes_excel(archivo):
    """Devuelve un diccionario {ID: Nombre} de la hoja Users."""
    agentes_dict = {}
    wb = openpyxl.load_workbook(archivo)
    if 'Users' in wb.sheetnames:
        ws_users = wb['Users']
        for row in ws_users.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                agentes_dict[str(row[0])] = str(row[1])
    return agentes_dict

def leer_ejemplos_excel(archivo):
    ejemplos = []
    if not os.path.exists(archivo):
        raise FileNotFoundError(f"No se encontró el archivo: {archivo}")
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    agentes_dict = leer_agentes_excel(archivo)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ejemplo = dict(zip(headers, row))
        if ejemplo.get('Título') and ejemplo.get('Descripción'):
            asignado = str(ejemplo.get('Asignado ID', ''))
            if asignado in agentes_dict:
                ejemplo['Agente'] = agentes_dict[asignado]
            elif asignado:
                ejemplo['Agente'] = asignado
            else:
                ejemplo['Agente'] = 'No asignado'
            ejemplos.append(ejemplo)
    if not ejemplos:
        raise ValueError("No se encontraron ejemplos válidos en el archivo.")
    return ejemplos

def normalizar_texto(texto):
    import string
    # Quitar tildes y pasar a minúsculas
    texto = ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn')
    texto = texto.lower()
    # Quitar puntuación
    texto = texto.translate(str.maketrans('', '', string.punctuation))
    return texto

STOPWORDS = set([
    'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas', 'de', 'del', 'y', 'o', 'en', 'a', 'por', 'para', 'con', 'sin', 'al', 'se', 'que', 'es', 'su', 'sus', 'lo', 'le', 'les', 'mi', 'tu', 'su', 'si', 'no', 'ya', 'pero', 'como', 'más', 'mas', 'muy', 'ha', 'han', 'fue', 'son', 'esta', 'este', 'estos', 'estas', 'esa', 'ese', 'esas', 'esos', 'hay', 'porque', 'cuando', 'donde', 'quien', 'cual', 'cuales', 'sobre', 'entre', 'desde', 'hasta', 'ni', 'tambien', 'tampoco', 'otro', 'otra', 'otros', 'otras', 'u', 'e', 'me', 'te', 'nos', 'vos', 'usted', 'ustedes', 'ellos', 'ellas', 'el', 'ella', 'al', 'del', 'su', 'sus', 'le', 'les', 'lo', 'la', 'las', 'los', 'se', 'si', 'ya', 'a', 'de', 'en', 'y', 'o', 'u', 'que', 'por', 'para', 'con', 'sin', 'es', 'un', 'una', 'unos', 'unas', 'mi', 'tu', 'su', 'nuestro', 'nuestra', 'nuestros', 'nuestras', 'vuestro', 'vuestra', 'vuestros', 'vuestras', 'este', 'esta', 'estos', 'estas', 'ese', 'esa', 'esos', 'esas', 'aquel', 'aquella', 'aquellos', 'aquellas', 'yo', 'tú', 'él', 'ella', 'nosotros', 'vosotros', 'ellos', 'ellas', 'usted', 'ustedes', 'mio', 'mia', 'mios', 'mias', 'tuyo', 'tuya', 'tuyos', 'tuyas', 'suyo', 'suya', 'suyos', 'suyas', 'nuestro', 'nuestra', 'nuestros', 'nuestras', 'vuestro', 'vuestra', 'vuestros', 'vuestras', 'este', 'esta', 'estos', 'estas', 'ese', 'esa', 'esos', 'esas', 'aquel', 'aquella', 'aquellos', 'aquellas', 'yo', 'tú', 'él', 'ella', 'nosotros', 'vosotros', 'ellos', 'ellas', 'usted', 'ustedes'
])

def seleccionar_ejemplos_relevantes(ejemplos, nuevo_chat, n=3):
    # Normalizar y quitar stopwords del chat
    palabras_chat = set([w for w in normalizar_texto(nuevo_chat).split() if w not in STOPWORDS])
    ejemplos_con_relevancia = []
    for ej in ejemplos:
        titulo = normalizar_texto(ej['Título'])
        descripcion = normalizar_texto(ej['Descripción'])
        palabras_titulo = set([w for w in titulo.split() if w not in STOPWORDS])
        palabras_desc = set([w for w in descripcion.split() if w not in STOPWORDS])
        # Ponderar coincidencias en el título x2
        relevancia = 2 * len(palabras_chat.intersection(palabras_titulo)) + len(palabras_chat.intersection(palabras_desc))
        ejemplos_con_relevancia.append((ej, relevancia))
    ejemplos_con_relevancia.sort(key=lambda x: x[1], reverse=True)
    ejemplos_seleccionados = []
    for ej, _ in ejemplos_con_relevancia[:n]:
        # Limitar la descripción a 40 palabras
        desc_corta = ' '.join(str(ej['Descripción']).split()[:40])
        ej_copia = ej.copy()
        ej_copia['Descripción'] = desc_corta + ('...' if len(str(ej['Descripción']).split()) > 40 else '')
        # Añadir el nombre del solicitante a la descripción
        solicitante = ej.get('Solicitante', 'Desconocido')
        ej_copia['Descripción'] = f"Solicitante: {solicitante}. {ej_copia['Descripción']}"
        ejemplos_seleccionados.append(ej_copia)
    # Si hay pocos relevantes, completa con aleatorios
    if len(ejemplos_seleccionados) < n:
        ejemplos_restantes = [ej[0] for ej in ejemplos_con_relevancia[n:]]
        if ejemplos_restantes:
            ejemplos_seleccionados.extend(random.sample(ejemplos_restantes, min(n - len(ejemplos_seleccionados), len(ejemplos_restantes))))
    return ejemplos_seleccionados

def construir_prompt(ejemplos, chat_usuario):
    agentes_unicos = sorted(set(ej['Agente'] for ej in ejemplos if ej['Agente'] and ej['Agente'] != 'No asignado'))
    prompt = (
        "Estos son ejemplos de tickets que yo mismo he escrito (asunto, descripción, agente asignado y solicitante). "
        "Analiza mi estilo y formato. Luego, a partir del siguiente chat, genera un ticket nuevo (asunto, descripción, agente recomendado y solicitante) imitando mi forma de escribir.\n"
        f"La lista de agentes disponibles es: {', '.join(agentes_unicos)}. Elige siempre uno de estos agentes como recomendado.\n"
        "IMPORTANTE: La descripción debe contener SOLO el problema reportado, NO incluyas la solución ni frases de cierre. La solución debe ir únicamente en el campo 'mensaje de cierre' que será agregado por el sistema.\n\n"
        "Ejemplos:\n"
    )
    for ej in ejemplos:
        prompt += f"- Solicitante: {ej.get('Solicitante', 'Desconocido')}\n  Asunto: {ej['Título']}\n  Descripción: {ej['Descripción']}\n  Agente asignado: {ej.get('Agente', '')}\n"
    prompt += (
        f"\nChat con usuario:\n{chat_usuario}\n\n"
        "Devuelve solo el ticket en formato JSON así: {\"solicitante\": \"...\", \"titulo\": \"...\", \"descripcion\": \"...\", \"agente\": \"...\"}"
    )
    return prompt

def extraer_json_de_texto(texto, agentes_validos=None):
    import json, re
    match = re.search(r'\{.*\}', texto, re.DOTALL)
    if match:
        try:
            data = json.loads(match.group(0))
            if agentes_validos is not None:
                if not data.get('agente') or data['agente'] not in agentes_validos:
                    data['agente'] = 'No asignado'
            # Si el campo solicitante está vacío o es desconocido, intentar extraerlo de la descripción
            if not data.get('solicitante') or data['solicitante'].lower() == 'desconocido':
                desc = data.get('descripcion', '')
                # Buscar patrones como 'El usuario NOMBRE ...' o 'El solicitante NOMBRE ...'
                m = re.search(r'(?:usuario|solicitante) ([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)', desc)
                if m:
                    data['solicitante'] = m.group(1)
            # Limpiar la descripción si contiene el nombre del solicitante al inicio
            if 'solicitante' in data and data['solicitante']:
                patron = re.compile(r'(Solicitante: ?' + re.escape(data['solicitante']) + r'\.? ?)', re.IGNORECASE)
                data['descripcion'] = patron.sub('', data.get('descripcion', ''), count=1).strip()
                # También quitar 'El usuario NOMBRE' al inicio
                data['descripcion'] = re.sub(r'^El usuario ' + re.escape(data['solicitante']) + r'\b[\.:,;\- ]*', '', data['descripcion'], flags=re.IGNORECASE)
            return data
        except Exception:
            pass
    # Fallback si no es JSON
    lineas = texto.split('\n')
    solicitante = ''
    titulo = ''
    descripcion = ''
    agente = ''
    for l in lineas:
        if 'Solicitante' in l or 'solicitante' in l:
            solicitante = l.split(':',1)[-1].strip()
        if 'Título:' in l or 'Asunto:' in l:
            titulo = l.split(':',1)[-1].strip()
        if 'Descripción:' in l:
            descripcion = l.split(':',1)[-1].strip()
        if 'Agente' in l or 'agente' in l:
            agente = l.split(':',1)[-1].strip()
    if agentes_validos is not None and (not agente or agente not in agentes_validos):
        agente = 'No asignado'
    # Intentar extraer solicitante de la descripción
    if not solicitante or solicitante.lower() == 'desconocido':
        m = re.search(r'(?:usuario|solicitante) ([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)', descripcion)
        if m:
            solicitante = m.group(1)
    # Limpiar la descripción si contiene el nombre del solicitante
    if solicitante:
        patron = re.compile(r'(Solicitante: ?' + re.escape(solicitante) + r'\.? ?)', re.IGNORECASE)
        descripcion = patron.sub('', descripcion, count=1).strip()
        descripcion = re.sub(r'^El usuario ' + re.escape(solicitante) + r'\b[\.:,;\- ]*', '', descripcion, flags=re.IGNORECASE)
    return {"solicitante": solicitante or 'Desconocido', "titulo": titulo or 'Error en formato', "descripcion": descripcion or texto, "agente": agente}

def pedir_a_gemini(prompt, status_callback=None):
    try:
        model = genai.GenerativeModel(MODEL_NAME)
        if status_callback:
            status_callback("Enviando prompt a Gemini...")
        response = model.generate_content(prompt)
        texto = response.text
        return extraer_json_de_texto(texto)
    except Exception as e:
        return {"titulo": "Error", "descripcion": f"Error al comunicarse con Gemini: {str(e)}"}

def guardar_ticket(ticket, archivo=DEFAULT_SAVE_FILE):
    try:
        with open(archivo, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([ticket['titulo'], ticket['descripcion'], 'Sistema', ''])
        return True
    except Exception as e:
        return False

# ===================== INTERFAZ GRÁFICA =====================
class TicketApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🎫 Generador de Tickets - Gemini (Pro)")
        self.root.geometry("950x750")
        self.root.configure(bg="#23272f")
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton', font=('Segoe UI', 12), background='#5865f2', foreground='white')
        style.configure('TLabel', font=('Segoe UI', 11), background='#23272f', foreground='white')
        style.configure('TEntry', font=('Segoe UI', 11))
        style.configure('TFrame', background='#23272f')

        # Atajos de teclado
        self.root.bind('<Control-n>', lambda e: self.nuevo_ticket())
        self.root.bind('<Control-N>', lambda e: self.nuevo_ticket())
        self.root.bind('<Control-s>', lambda e: self.guardar_ticket())
        self.root.bind('<Control-S>', lambda e: self.guardar_ticket())
        self.root.bind('<Control-f>', lambda e: self.historial_buscar_entry.focus_set())
        self.root.bind('<Control-F>', lambda e: self.historial_buscar_entry.focus_set())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        self.root.bind('<Control-Q>', lambda e: self.root.quit())
        # Deshacer/rehacer en campos de texto
        self.root.bind_all('<Control-z>', lambda e: self._deshacer())
        self.root.bind_all('<Control-y>', lambda e: self._rehacer())

        # Variables de archivos y parámetros
        self.excel_file = tk.StringVar(value=DEFAULT_EXCEL_FILE)
        self.save_file = tk.StringVar(value=DEFAULT_SAVE_FILE)
        self.num_ejemplos = tk.IntVar(value=DEFAULT_NUM_EJEMPLOS)
        self.prompt_ultimo = ""
        self.agentes_validos = []

        # Validar API y modelo
        valido, msg = validar_api_key_y_modelo()
        if not valido:
            messagebox.showerror("Error de API/Modelo", f"No se pudo validar la API Key o el modelo:\n{msg}")
            root.destroy()
            return

        # Cargar ejemplos
        try:
            self.ejemplos = leer_ejemplos_excel(self.excel_file.get())
            self.agentes_validos = sorted(set(ej['Agente'] for ej in self.ejemplos if ej['Agente'] and ej['Agente'] != 'No asignado'))
        except Exception as e:
            messagebox.showerror("Error al leer ejemplos", str(e))
            root.destroy()
            return

        # --- UI ---
        main_frame = ttk.Frame(root, padding=20)
        main_frame.pack(fill='both', expand=True)

        # Selección de archivo de ejemplos
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill='x', pady=(0,10))
        ttk.Label(file_frame, text="Base de ejemplos:").pack(side='left')
        ttk.Entry(file_frame, textvariable=self.excel_file, width=40, state='readonly').pack(side='left', padx=5)
        ttk.Button(file_frame, text="Cambiar...", command=self.cambiar_excel).pack(side='left', padx=5)
        ttk.Label(file_frame, text="Ejemplos a usar:").pack(side='left', padx=(20,0))
        ttk.Spinbox(file_frame, from_=1, to=20, textvariable=self.num_ejemplos, width=3).pack(side='left', padx=5)

        # --- Panel de historial de tickets ---
        historial_frame = tk.Frame(main_frame, bg='#23272f')
        historial_frame.pack(fill='both', expand=False, pady=(10, 10))
        tk.Label(historial_frame, text="Historial de Tickets Guardados", bg='#23272f', fg='white', font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0,5), padx=2)
        buscar_frame = tk.Frame(historial_frame, bg='#23272f')
        buscar_frame.pack(fill='x', pady=(0,5))
        tk.Label(buscar_frame, text="Buscar por asunto:", bg='#23272f', fg='white', font=('Segoe UI', 11)).pack(side='left')
        self.historial_buscar_var = tk.StringVar()
        self.historial_buscar_entry = tk.Entry(buscar_frame, textvariable=self.historial_buscar_var, width=30, bg='#181a20', fg='white', insertbackground='white', relief='flat', font=('Segoe UI', 11))
        self.historial_buscar_entry.pack(side='left', padx=5)
        self.historial_buscar_entry.bind('<KeyRelease>', lambda e: self.actualizar_historial())
        self.historial_listbox = tk.Listbox(historial_frame, width=100, height=5, bg="#181a20", fg="white", selectbackground="#5865f2", selectforeground="white", highlightbackground="#23272f", highlightcolor="#5865f2", relief='flat', borderwidth=0, font=('Consolas', 11))
        self.historial_listbox.pack(fill='x', pady=(0,5), padx=2)
        tk.Button(historial_frame, text="Abrir Ticket", command=self.abrir_ticket_historial, bg='#5865f2', fg='white', activebackground='#7289da', activeforeground='white', relief='flat', font=('Segoe UI', 11, 'bold')).pack(side='right', padx=2, pady=2)
        self.actualizar_historial()

        # --- Separador visual ---
        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=10)

        # Chat input
        ttk.Label(main_frame, text="Pega aquí el chat completo con el usuario:").pack(anchor='w', pady=(0,5))
        self.chat_text = scrolledtext.ScrolledText(main_frame, width=100, height=8, font=('Consolas', 11), bg="#181a20", fg="white", insertbackground="white", undo=True, maxundo=-1)
        self.chat_text.pack(pady=5, fill='x')

        # Botón generar
        self.generar_btn = ttk.Button(main_frame, text="✨ Generar Ticket", command=self.generar_ticket)
        self.generar_btn.pack(pady=10, fill='x')

        # Campo Solicitante (debe ir arriba de Asunto generado)
        solicitante_frame = ttk.Frame(main_frame)
        solicitante_frame.pack(fill='x', pady=(10,0))
        ttk.Label(solicitante_frame, text="Solicitante:").pack(side='left')
        self.solicitante_var = tk.StringVar()
        self.solicitante_entry = ttk.Entry(solicitante_frame, textvariable=self.solicitante_var, width=80, font=('Segoe UI', 11), state='readonly')
        self.solicitante_entry.pack(side='left', padx=5, fill='x', expand=True)

        # Asunto y copiar
        asunto_frame = ttk.Frame(main_frame)
        asunto_frame.pack(fill='x', pady=(10,0))
        ttk.Label(asunto_frame, text="Asunto generado:").pack(side='left')
        self.asunto_var = tk.StringVar()
        self.asunto_entry = ttk.Entry(asunto_frame, textvariable=self.asunto_var, width=80, font=('Segoe UI', 11), state='readonly')
        self.asunto_entry.pack(side='left', padx=5, fill='x', expand=True)
        ttk.Button(asunto_frame, text="Copiar", command=self.copiar_asunto).pack(side='left', padx=5)

        # Descripción y copiar
        desc_frame = ttk.Frame(main_frame)
        desc_frame.pack(fill='x', pady=(10,0))
        ttk.Label(desc_frame, text="Descripción generada:").pack(side='left')
        self.desc_text = scrolledtext.ScrolledText(desc_frame, width=100, height=6, font=('Consolas', 11), bg="#181a20", fg="white", insertbackground="white", undo=True, maxundo=-1)
        self.desc_text.pack(side='left', padx=5, fill='x', expand=True)
        self.desc_text.config(state='disabled')
        ttk.Button(desc_frame, text="Copiar", command=self.copiar_descripcion).pack(side='left', padx=5)

        # Checkbox para marcar como solucionado
        self.solucionado_var = tk.BooleanVar()
        self.solucionado_check = ttk.Checkbutton(main_frame, text="Marcar como solucionado", variable=self.solucionado_var, command=self.mostrar_mensaje_cierre)
        self.solucionado_check.pack(anchor='w', pady=(5,0))

        # Campo de mensaje de cierre (solo visible si está marcado)
        cierre_frame = ttk.Frame(main_frame)
        cierre_frame.pack(fill='x', pady=(0,0))
        self.cierre_frame = cierre_frame
        self.mensaje_cierre_label = ttk.Label(cierre_frame, text="Mensaje de cierre:")
        self.mensaje_cierre_entry = ttk.Entry(cierre_frame, state='readonly', font=('Segoe UI', 11))
        self.copiar_cierre_btn = ttk.Button(cierre_frame, text="Copiar", command=self.copiar_mensaje_cierre)
        # Inicialmente ocultos
        self.mensaje_cierre_label.pack_forget()
        self.mensaje_cierre_entry.pack_forget()
        self.copiar_cierre_btn.pack_forget()

        # Agente recomendado
        ttk.Label(main_frame, text="Agente recomendado:").pack(anchor='w', pady=(10,0))
        self.agente_var = tk.StringVar()
        self.agente_entry = ttk.Entry(main_frame, textvariable=self.agente_var, width=80, font=('Segoe UI', 11), state='readonly')
        self.agente_entry.pack(pady=2, fill='x')

        # Botón ver prompt
        self.ver_prompt_btn = ttk.Button(main_frame, text="Ver Prompt Enviado", command=self.ver_prompt)
        self.ver_prompt_btn.pack(pady=(5,0), fill='x')

        # Guardar, Nuevo, Salir
        botones_frame = ttk.Frame(main_frame)
        botones_frame.pack(fill='x', pady=(20,0))
        self.guardar_btn = ttk.Button(botones_frame, text="💾 Guardar Ticket", command=self.guardar_ticket, state='disabled')
        self.guardar_btn.pack(side='left', expand=True, fill='x', padx=(0,10))
        self.nuevo_btn = ttk.Button(botones_frame, text="🆕 Nuevo Ticket", command=self.nuevo_ticket)
        self.nuevo_btn.pack(side='left', expand=True, fill='x', padx=(0,10))
        self.salir_btn = ttk.Button(botones_frame, text="❌ Salir", command=self.root.quit)
        self.salir_btn.pack(side='right', expand=True, fill='x', padx=(10,0))

        # Barra de estado
        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(main_frame, textvariable=self.status_var, font=('Segoe UI', 10, 'italic'))
        self.status_label.pack(anchor='w', pady=(10,0))

    def cambiar_excel(self):
        file = filedialog.askopenfilename(title="Selecciona archivo de ejemplos", filetypes=[("Excel files", "*.xlsx")])
        if file:
            try:
                ejemplos = leer_ejemplos_excel(file)
                self.ejemplos = ejemplos
                self.excel_file.set(file)
                self.agentes_validos = sorted(set(ej['Agente'] for ej in self.ejemplos if ej['Agente'] and ej['Agente'] != 'No asignado'))
                self.status_var.set(f"Base de ejemplos cargada: {os.path.basename(file)}")
            except Exception as e:
                messagebox.showerror("Error al leer ejemplos", str(e))

    def generar_ticket(self):
        chat = self.chat_text.get("1.0", tk.END).strip()
        if not chat:
            messagebox.showwarning("Campos vacíos", "Por favor, pega el chat con el usuario.")
            return
        self.status_var.set("Generando ticket...")
        self.generar_btn.config(state='disabled')
        self.root.update_idletasks()
        threading.Thread(target=self._generar_ticket_thread, args=(chat,)).start()

    def _generar_ticket_thread(self, chat):
        try:
            ejemplos_relevantes = seleccionar_ejemplos_relevantes(self.ejemplos, chat, n=self.num_ejemplos.get())
            prompt = construir_prompt(ejemplos_relevantes, chat)
            self.prompt_ultimo = prompt
            ticket = pedir_a_gemini(prompt, status_callback=self.status_var.set)
            self.agentes_validos = sorted(set(ej['Agente'] for ej in self.ejemplos if ej['Agente'] and ej['Agente'] != 'No asignado'))
            ticket = extraer_json_de_texto(texto=ticket if isinstance(ticket, str) else ticket['descripcion'], agentes_validos=self.agentes_validos) if isinstance(ticket, dict) and 'descripcion' in ticket and isinstance(ticket['descripcion'], str) and ticket['descripcion'].startswith('{') else ticket
            self.root.after(0, self._mostrar_ticket, ticket)
        except Exception as e:
            self.root.after(0, self._mostrar_ticket, {"titulo": "Error", "descripcion": str(e)})

    def _mostrar_ticket(self, ticket):
        self.asunto_var.set(ticket['titulo'])
        descripcion = ticket.get('descripcion', '')
        # Eliminar cualquier aparición de 'Solicitante: ...' en la descripción
        import re
        descripcion = re.sub(r'Solicitante: ?[^\.\n]*[\.\n]?', '', descripcion, flags=re.IGNORECASE).strip()
        # Eliminar frases típicas de solución si aparecen en la descripción
        frases_solucion = [
            'quedó solucionado', 'incidente resuelto', 'incidente cerrado', 'problema resuelto',
            'fue solucionado', 'ya está solucionado', 'ya fue resuelto', 'ya fue solucionado',
            'se resolvió', 'se solucionó', 'solucionado', 'resuelto', 'cerrado', 'quedó resuelto'
        ]
        for frase in frases_solucion:
            patron = re.compile(frase, re.IGNORECASE)
            descripcion = patron.sub('', descripcion)
        descripcion = re.sub(r'Estimad[oa]s?,? ?qued[óo] solucionado el incidente\. Saludos\.?', '', descripcion, flags=re.IGNORECASE)
        self.desc_text.config(state='normal')
        self.desc_text.delete("1.0", tk.END)
        self.desc_text.insert(tk.END, descripcion.strip())
        self.desc_text.config(state='disabled')
        self.agente_var.set(ticket.get('agente', ''))
        self.solicitante_var.set(ticket.get('solicitante', 'Desconocido'))
        # Mostrar mensaje de cierre si está marcado como solucionado
        self.mostrar_mensaje_cierre()
        if ticket.get('agente', '') == 'No asignado':
            self.agente_entry.config(foreground='#f04747')
            self.status_var.set("Advertencia: El agente recomendado no está en la lista de agentes válidos.")
            self.status_label.config(foreground='#f04747')
        else:
            self.agente_entry.config(foreground='black')
        if ticket.get('titulo') not in ['Error', 'Error en formato', 'Error de tamaño']:
            self.guardar_btn.config(state='normal')
            self.status_var.set("Ticket generado correctamente.")
            self.status_label.config(foreground='#43b581')
        else:
            self.guardar_btn.config(state='disabled')
            self.status_var.set(f"Error al generar el ticket: {ticket['descripcion']}")
            self.status_label.config(foreground='#f04747')
        self.generar_btn.config(state='normal')

    def guardar_ticket(self, *_):
        ticket = {
            'titulo': self.asunto_var.get(),
            'descripcion': self.desc_text.get("1.0", tk.END).strip()
        }
        if not ticket['titulo'].strip() or not ticket['descripcion'].strip():
            messagebox.showwarning("Campos vacíos", "El asunto y la descripción no pueden estar vacíos.")
            return
        carpeta = os.path.join(os.path.dirname(__file__), 'tickets_guardados')
        os.makedirs(carpeta, exist_ok=True)
        nombre_archivo = re.sub(r'[^\w\- ]', '', ticket['titulo'])[:50].strip().replace(' ', '_')
        if not nombre_archivo:
            nombre_archivo = 'ticket_sin_asunto'
        ruta = os.path.join(carpeta, f"{nombre_archivo}.txt")
        if os.path.exists(ruta):
            messagebox.showwarning("Duplicado", f"Ya existe un ticket con ese asunto. Cambia el asunto para guardar otro ticket.")
            return
        mensaje_cierre = ""
        if self.solucionado_var.get():
            mensaje_cierre = self.mensaje_cierre_entry.get()
        try:
            with open(ruta, 'w', encoding='utf-8') as f:
                f.write(f"Asunto: {ticket['titulo']}\n\nDescripción (problema):\n{ticket['descripcion']}\n")
                if mensaje_cierre:
                    f.write(f"\nMensaje de cierre:\n{mensaje_cierre}\n")
            messagebox.showinfo("Éxito", f"Ticket guardado como archivo:\n{ruta}")
            self.status_var.set(f"Ticket guardado en {ruta}.")
            self.status_label.config(foreground='#43b581')
            self.actualizar_historial()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el ticket como archivo:\n{e}")
            self.status_var.set("No se pudo guardar el ticket como archivo.")
            self.status_label.config(foreground='#f04747')

    def nuevo_ticket(self, *_):
        self.chat_text.delete("1.0", tk.END)
        self.asunto_var.set("")
        self.desc_text.config(state='normal')
        self.desc_text.delete("1.0", tk.END)
        self.desc_text.config(state='disabled')
        self.guardar_btn.config(state='disabled')
        self.status_var.set("")
        self.status_label.config(foreground='white')
        self.agente_var.set("")
        self.solicitante_var.set("")
        self.solucionado_var.set(False)
        self.mostrar_mensaje_cierre()

    def copiar_asunto(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.asunto_var.get())
        self.status_var.set("Asunto copiado al portapapeles.")
        self.status_label.config(foreground='#7289da')

    def copiar_descripcion(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.desc_text.get("1.0", tk.END).strip())
        self.status_var.set("Descripción copiada al portapapeles.")
        self.status_label.config(foreground='#7289da')

    def ver_prompt(self):
        top = tk.Toplevel(self.root)
        top.title("Prompt enviado a Gemini")
        top.geometry("800x400")
        text = scrolledtext.ScrolledText(top, width=100, height=20, font=('Consolas', 10))
        text.pack(fill='both', expand=True)
        text.insert(tk.END, self.prompt_ultimo)
        text.config(state='disabled')
        ttk.Button(top, text="Cerrar", command=top.destroy).pack(pady=5)

    def actualizar_historial(self):
        carpeta = os.path.join(os.path.dirname(__file__), 'tickets_guardados')
        patron = os.path.join(carpeta, '*.txt')
        archivos = glob.glob(patron)
        filtro = self.historial_buscar_var.get().lower()
        self.historial_listbox.delete(0, tk.END)
        for archivo in archivos:
            nombre = os.path.splitext(os.path.basename(archivo))[0]
            if filtro in nombre.lower():
                self.historial_listbox.insert(tk.END, nombre)

    def abrir_ticket_historial(self):
        seleccion = self.historial_listbox.curselection()
        if not seleccion:
            messagebox.showinfo("Selecciona un ticket", "Selecciona un ticket de la lista para abrirlo.")
            return
        nombre = self.historial_listbox.get(seleccion[0])
        carpeta = os.path.join(os.path.dirname(__file__), 'tickets_guardados')
        ruta = os.path.join(carpeta, f"{nombre}.txt")
        if not os.path.exists(ruta):
            messagebox.showerror("No encontrado", "No se encontró el archivo del ticket seleccionado.")
            return
        with open(ruta, 'r', encoding='utf-8') as f:
            contenido = f.read()
        top = tk.Toplevel(self.root)
        top.title(f"Ticket: {nombre}")
        text = scrolledtext.ScrolledText(top, width=80, height=20, font=('Consolas', 11))
        text.pack(fill='both', expand=True)
        text.insert(tk.END, contenido)
        text.config(state='disabled')
        ttk.Button(top, text="Cerrar", command=top.destroy).pack(pady=5)

    def _deshacer(self, *_):
        try:
            widget = self.root.focus_get()
            widget.edit_undo()
        except Exception:
            pass

    def _rehacer(self, *_):
        try:
            widget = self.root.focus_get()
            widget.edit_redo()
        except Exception:
            pass

    def mostrar_mensaje_cierre(self):
        if self.solucionado_var.get():
            self.mensaje_cierre_label.pack(side='left', padx=(0,5))
            self.mensaje_cierre_entry.config(state='normal')
            self.mensaje_cierre_entry.delete(0, tk.END)
            self.mensaje_cierre_entry.insert(0, "Estimados, quedó solucionado el incidente. Saludos.")
            self.mensaje_cierre_entry.config(state='readonly')
            self.mensaje_cierre_entry.pack(side='left', fill='x', expand=True)
            self.copiar_cierre_btn.pack(side='left', padx=5)
        else:
            self.mensaje_cierre_label.pack_forget()
            self.mensaje_cierre_entry.pack_forget()
            self.copiar_cierre_btn.pack_forget()

    def copiar_mensaje_cierre(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.mensaje_cierre_entry.get())
        self.status_var.set("Mensaje de cierre copiado al portapapeles.")
        self.status_label.config(foreground='#7289da')

if __name__ == "__main__":
    root = tk.Tk()
    app = TicketApp(root)
    root.mainloop() 