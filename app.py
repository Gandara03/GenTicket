import os
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash
import openpyxl
from werkzeug.utils import secure_filename
import google.generativeai as genai
import random
import re
import unicodedata
import json

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'tickets_guardados')
EXCEL_DEFAULT = 'Tickets_1.xlsx'
ALLOWED_EXTENSIONS = {'xlsx'}
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
MODEL_NAME = 'gemini-2.0-flash'

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'supersecretkey')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_historial():
    carpeta = UPLOAD_FOLDER
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)
    archivos = [f for f in os.listdir(carpeta) if f.endswith('.txt')]
    return sorted(archivos, reverse=True)

def get_excel_files():
    files = [f for f in os.listdir(os.path.dirname(__file__)) if f.endswith('.xlsx')]
    return sorted(files)

def normalizar_texto(texto):
    texto = ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn')
    texto = texto.lower()
    texto = texto.translate(str.maketrans('', '', '.,;:!?()[]{}"'))
    return texto

STOPWORDS = set([
    'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas', 'de', 'del', 'y', 'o', 'en', 'a', 'por', 'para', 'con', 'sin', 'al', 'se', 'que', 'es', 'su', 'sus', 'lo', 'le', 'les', 'mi', 'tu', 'su', 'si', 'no', 'ya', 'pero', 'como', 'más', 'mas', 'muy', 'ha', 'han', 'fue', 'son', 'esta', 'este', 'estos', 'estas', 'esa', 'ese', 'esas', 'esos', 'hay', 'porque', 'cuando', 'donde', 'quien', 'cual', 'cuales', 'sobre', 'entre', 'desde', 'hasta', 'ni', 'tambien', 'tampoco', 'otro', 'otra', 'otros', 'otras', 'u', 'e', 'me', 'te', 'nos', 'vos', 'usted', 'ustedes', 'ellos', 'ellas', 'el', 'ella', 'al', 'del', 'su', 'sus', 'le', 'les', 'lo', 'la', 'las', 'los', 'se', 'si', 'ya', 'a', 'de', 'en', 'y', 'o', 'u', 'que', 'por', 'para', 'con', 'sin', 'es', 'un', 'una', 'unos', 'unas', 'mi', 'tu', 'su', 'nuestro', 'nuestra', 'nuestros', 'nuestras', 'vuestro', 'vuestra', 'vuestros', 'vuestras', 'este', 'esta', 'estos', 'estas', 'ese', 'esa', 'esos', 'esas', 'aquel', 'aquella', 'aquellos', 'aquellas', 'yo', 'tú', 'él', 'ella', 'nosotros', 'vosotros', 'ellos', 'ellas', 'usted', 'ustedes', 'mio', 'mia', 'mios', 'mias', 'tuyo', 'tuya', 'tuyos', 'tuyas', 'suyo', 'suya', 'suyos', 'suyas', 'nuestro', 'nuestra', 'nuestros', 'nuestras', 'vuestro', 'vuestra', 'vuestros', 'vuestras', 'este', 'esta', 'estos', 'estas', 'ese', 'esa', 'esos', 'esas', 'aquel', 'aquella', 'aquellos', 'aquellas', 'yo', 'tú', 'él', 'ella', 'nosotros', 'vosotros', 'ellos', 'ellas', 'usted', 'ustedes'
])

def leer_agentes_excel(archivo):
    agentes_dict = {}
    wb = openpyxl.load_workbook(archivo)
    if 'Users' in wb.sheetnames:
        ws_users = wb['Users']
        for row in ws_users.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                agentes_dict[str(row[0])] = str(row[1])
    return agentes_dict

def leer_ejemplos_excel(archivo):
    ejemplos = []
    if not os.path.exists(archivo):
        return ejemplos
    try:
        wb = openpyxl.load_workbook(archivo)
    except Exception:
        return ejemplos
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    agentes_dict = leer_agentes_excel(archivo)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ejemplo = dict(zip(headers, row))
        if ejemplo.get('Título') and ejemplo.get('Descripción'):
            asignado = str(ejemplo.get('Asignado ID', ''))
            if asignado in agentes_dict:
                ejemplo['Agente'] = agentes_dict[asignado]
            elif asignado:
                ejemplo['Agente'] = asignado
            else:
                ejemplo['Agente'] = 'No asignado'
            ejemplos.append(ejemplo)
    return ejemplos

def seleccionar_ejemplos_relevantes(ejemplos, nuevo_chat, n=3):
    palabras_chat = set([w for w in normalizar_texto(nuevo_chat).split() if w not in STOPWORDS])
    ejemplos_con_relevancia = []
    for ej in ejemplos:
        titulo = normalizar_texto(ej['Título'])
        descripcion = normalizar_texto(ej['Descripción'])
        palabras_titulo = set([w for w in titulo.split() if w not in STOPWORDS])
        palabras_desc = set([w for w in descripcion.split() if w not in STOPWORDS])
        relevancia = 2 * len(palabras_chat.intersection(palabras_titulo)) + len(palabras_chat.intersection(palabras_desc))
        ejemplos_con_relevancia.append((ej, relevancia))
    ejemplos_con_relevancia.sort(key=lambda x: x[1], reverse=True)
    ejemplos_seleccionados = []
    for ej, _ in ejemplos_con_relevancia[:n]:
        desc_corta = ' '.join(str(ej['Descripción']).split()[:40])
        ej_copia = ej.copy()
        ej_copia['Descripción'] = desc_corta + ('...' if len(str(ej['Descripción']).split()) > 40 else '')
        solicitante = ej.get('Solicitante', 'Desconocido')
        ej_copia['Descripción'] = f"Solicitante: {solicitante}. {ej_copia['Descripción']}"
        ejemplos_seleccionados.append(ej_copia)
    if len(ejemplos_seleccionados) < n:
        ejemplos_restantes = [ej[0] for ej in ejemplos_con_relevancia[n:]]
        if ejemplos_restantes:
            ejemplos_seleccionados.extend(random.sample(ejemplos_restantes, min(n - len(ejemplos_seleccionados), len(ejemplos_restantes))))
    return ejemplos_seleccionados

def construir_prompt(ejemplos, chat_usuario):
    agentes_unicos = sorted(set(ej['Agente'] for ej in ejemplos if ej['Agente'] and ej['Agente'] != 'No asignado'))
    prompt = (
        "Estos son ejemplos de tickets que yo mismo he escrito (asunto, descripción, agente asignado y solicitante). "
        "Analiza mi estilo y formato. Luego, a partir del siguiente chat, genera un ticket nuevo (asunto, descripción, agente recomendado y solicitante) imitando mi forma de escribir.\n"
        f"La lista de agentes disponibles es: {', '.join(agentes_unicos)}. Elige siempre uno de estos agentes como recomendado.\n"
        "IMPORTANTE: La descripción debe contener SOLO el problema reportado, NO incluyas la solución ni frases de cierre. La solución debe ir únicamente en el campo 'mensaje de cierre' que será agregado por el sistema.\n\n"
        "Ejemplos:\n"
    )
    for ej in ejemplos:
        prompt += f"- Solicitante: {ej.get('Solicitante', 'Desconocido')}\n  Asunto: {ej['Título']}\n  Descripción: {ej['Descripción']}\n  Agente asignado: {ej.get('Agente', '')}\n"
    prompt += (
        f"\nChat con usuario:\n{chat_usuario}\n\n"
        "IMPORTANTE: El campo 'solicitante' debe ser el nombre de la persona que inicia el chat o hace la consulta. Si no está claro, dedúcelo del chat, pero nunca pongas 'Desconocido'. "
        "Devuelve solo el ticket en formato JSON así: {\"solicitante\": \"...\", \"titulo\": \"...\", \"descripcion\": \"...\", \"agente\": \"...\"}"
    )
    return prompt

def extraer_json_de_texto(texto, agentes_validos=None):
    match = re.search(r'\{.*\}', texto, re.DOTALL)
    if match:
        try:
            data = json.loads(match.group(0))
            if agentes_validos is not None:
                if not data.get('agente') or data['agente'] not in agentes_validos:
                    data['agente'] = 'No asignado'
            # Mejorar la detección del solicitante
            if not data.get('solicitante') or data['solicitante'].lower() == 'desconocido':
                desc = data.get('descripcion', '')
                patrones = [
                    r'^([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+):',
                    r'(?:usuario|solicitante|cliente) ([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)',
                    r'^([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)\s*:',
                    r'^([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)\s*$',
                    # Nuevo patrón para nombre seguido de emoji y hora
                    r'^([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?: [A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)\s*[:\w\W]*\d{1,2}:\d{2}'
                ]
                for patron in patrones:
                    m = re.search(patron, desc, re.MULTILINE)
                    if m:
                        data['solicitante'] = m.group(1)
                        break
            # Limpiar la descripción de referencias al solicitante
            if 'solicitante' in data and data['solicitante']:
                patron = re.compile(r'(Solicitante: ?' + re.escape(data['solicitante']) + r'\.? ?)', re.IGNORECASE)
                data['descripcion'] = patron.sub('', data.get('descripcion', ''), count=1).strip()
                data['descripcion'] = re.sub(r'^El usuario ' + re.escape(data['solicitante']) + r'\b[\.:,;\- ]*', '', data['descripcion'], flags=re.IGNORECASE)
            return data
        except Exception:
            pass
    return {"solicitante": 'Desconocido', "titulo": 'Error', "descripcion": texto, "agente": ''}

def pedir_a_gemini(prompt):
    if not GEMINI_API_KEY:
        return {"solicitante": 'Desconocido', "titulo": 'Error', "descripcion": 'No se configuró la API KEY de Gemini.', "agente": ''}
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel(MODEL_NAME)
        response = model.generate_content(prompt)
        texto = response.text
        return extraer_json_de_texto(texto)
    except Exception as e:
        return {"solicitante": 'Desconocido', "titulo": 'Error', "descripcion": str(e), "agente": ''}

@app.route('/', methods=['GET', 'POST'])
def index():
    asunto = descripcion = solicitante = agente = mensaje_cierre = ''
    historial = get_historial()
    excel_files = get_excel_files()
    excel_file = request.form.get('excel_file', EXCEL_DEFAULT)
    if excel_file not in excel_files:
        excel_file = EXCEL_DEFAULT
    if request.method == 'POST':
        chat = request.form.get('chat', '')[:5000]
        num_ejemplos = int(request.form.get('num_ejemplos', 3))
        solucionado = request.form.get('solucionado') == 'on'
        ejemplos = leer_ejemplos_excel(excel_file)
        if not ejemplos:
            flash('No se encontraron ejemplos válidos en el Excel.', 'danger')
            return redirect(url_for('index'))
        ejemplos_relevantes = seleccionar_ejemplos_relevantes(ejemplos, chat, n=num_ejemplos)
        prompt = construir_prompt(ejemplos_relevantes, chat)
        ticket = pedir_a_gemini(prompt)
        asunto = ticket.get('titulo', '')[:200]
        descripcion = ticket.get('descripcion', '')[:2000]
        solicitante = ticket.get('solicitante', '')[:100]
        agente = ticket.get('agente', '')[:100]
        mensaje_cierre = 'Estimados, quedó solucionado el incidente. Saludos.' if solucionado else ''
        if 'guardar' in request.form and asunto:
            nombre_archivo = secure_filename(asunto)[:50] or 'ticket_sin_asunto'
            ruta = os.path.join(UPLOAD_FOLDER, f"{nombre_archivo}.txt")
            try:
                with open(ruta, 'w', encoding='utf-8') as f:
                    f.write(f"Asunto: {asunto}\n\nDescripción (problema):\n{descripcion}\n")
                    if mensaje_cierre:
                        f.write(f"\nMensaje de cierre:\n{mensaje_cierre}\n")
                flash(f"Ticket guardado como {nombre_archivo}.txt", 'success')
            except Exception as e:
                flash(f"Error al guardar el ticket: {e}", 'danger')
            return redirect(url_for('index'))
    return render_template('index.html',
        asunto=asunto,
        descripcion=descripcion,
        solicitante=solicitante,
        agente=agente,
        mensaje_cierre=mensaje_cierre,
        historial=historial,
        excel_default=excel_file,
        excel_files=get_excel_files()
    )

@app.route('/tickets/<filename>')
def download_ticket(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.run(debug=True, host='0.0.0.0', port=5000) 